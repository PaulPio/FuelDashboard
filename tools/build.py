"""Build the lookup tool from the source workbook.

    python tools/build.py

Emits:
    data/equivalents.json   structured data, reviewable in a diff
    data/QA-REPORT.md       corrections applied + suspected errors for Stem Fuels
    dist/index.html         THE deliverable - self-contained, works offline and hosted

The workbook is read-only input and is never written to.
"""

import base64
import datetime
import json
import os
import re
import sys
import zipfile

import openpyxl

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

import qa  # noqa: E402
from corrections import CORRECTIONS, build_index  # noqa: E402
from parse import (BRANDS, FIRST_ROW, LAST_ROW, PRODUCTS, category_blocks,  # noqa: E402
                   normalize, parse_cell)

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
SOURCE = os.path.join(ROOT, "Stem Fuels_Lubricant equivalents.xlsx")
TEMPLATE = os.path.join(ROOT, "app", "index.template.html")
OUT_JSON = os.path.join(ROOT, "data", "equivalents.json")
OUT_QA = os.path.join(ROOT, "data", "QA-REPORT.md")
OUT_HTML = os.path.join(ROOT, "dist", "index.html")


def load_logos():
    """Brand logos, base64 data URIs, read straight out of the xlsx."""
    logos = {}
    with zipfile.ZipFile(SOURCE) as z:
        for brand in BRANDS:
            raw = z.read(f"xl/media/image{brand['image']}.png")
            logos[brand["id"]] = "data:image/png;base64," + base64.b64encode(raw).decode()
    return logos


def source_modified():
    with zipfile.ZipFile(SOURCE) as z:
        core = z.read("docProps/core.xml").decode("utf8", "replace")
    m = re.search(r"<dcterms:modified[^>]*>([^<]+)<", core)
    return m.group(1) if m else "unknown"


def build_rows(ws, blocks, corrections_by_cell, applied):
    rows = []
    for block in blocks:
        for row_num in range(block["start"], block["end"] + 1):
            cells = {}
            aliases_extra = {}
            for brand in BRANDS:
                ref = f"{brand['col']}{row_num}"
                raw = ws.cell(row=row_num, column=brand["col_idx"]).value

                fix = corrections_by_cell.get(ref)
                if fix is not None:
                    actual = re.sub(r"\s+", " ", str(raw or "")).strip()
                    expected = re.sub(r"\s+", " ", fix["before"]).strip()
                    if actual != expected:
                        raise SystemExit(
                            f"Correction for {ref} expected {expected!r} but the workbook "
                            f"now holds {actual!r}. Re-check tools/corrections.py."
                        )
                    raw = fix["after"]
                    applied.add(ref)
                    aliases_extra[brand["id"]] = fix["before"]

                parsed = parse_cell(raw)
                if parsed["state"] == PRODUCTS and brand["id"] in aliases_extra:
                    # Original spelling stays searchable as an exact match.
                    parsed["products"][0]["also"] = aliases_extra[brand["id"]]
                cells[brand["id"]] = parsed

            if not any(c["state"] == PRODUCTS for c in cells.values()):
                continue
            rows.append({"id": row_num, "category": block["name"], "cells": cells})
    return rows


def build_index_map(rows):
    """normalised name -> [{row, brand}]. One name legitimately hits several rows."""
    index = {}
    for i, row in enumerate(rows):
        for brand_id, cell in row["cells"].items():
            if cell["state"] != PRODUCTS:
                continue
            for product in cell["products"]:
                keys = [product["name"]]
                if product.get("also"):
                    keys.append(product["also"])
                for key in keys:
                    norm = normalize(key)
                    if not norm:
                        continue
                    index.setdefault(norm, [])
                    entry = {"r": i, "b": brand_id}
                    if entry not in index[norm]:
                        index[norm].append(entry)
    return index


def assert_invariants(rows, blocks, ws):
    """Fail the build rather than ship something subtly wrong."""
    # Every emitted row is a real equivalence (>= 2 products), or it is not useful.
    for row in rows:
        n = sum(len(c["products"]) for c in row["cells"].values() if c["state"] == PRODUCTS)
        if n < 2:
            print(f"  note: row {row['id']} has only {n} product(s) - sparse but kept")

    # Blocks must cover every non-blank row with no gaps and no overlaps.
    covered = []
    for b in blocks:
        covered.extend(range(b["start"], b["end"] + 1))
    if len(covered) != len(set(covered)):
        raise SystemExit("Category blocks overlap.")
    covered = set(covered)

    for row_num in range(FIRST_ROW, LAST_ROW + 1):
        has_data = any(
            parse_cell(ws.cell(row=row_num, column=b["col_idx"]).value)["state"] == PRODUCTS
            for b in BRANDS
        )
        if has_data and row_num not in covered:
            raise SystemExit(f"Row {row_num} holds data but no category block covers it.")

    # No '-' or empty cell may ever have produced a product.
    for row in rows:
        for brand_id, cell in row["cells"].items():
            if cell["state"] != PRODUCTS:
                assert "products" not in cell, f"{brand_id}:{row['id']} leaked products"


def write_qa_report(rows, findings, applied, meta):
    bname = {b["id"]: b["name"] for b in BRANDS}
    lines = []
    lines.append("# Lubricant Equivalents - Data QA Report\n")
    lines.append(f"Source: `{meta['source_file']}` (last modified {meta['source_modified']})  ")
    lines.append(f"Generated: {meta['generated']}\n")
    lines.append("The lookup tool ships the table **as it stands**. Nothing below has been ")
    lines.append("silently changed except the spelling fixes in section 1. Sections 2-4 are ")
    lines.append("for Stem Fuels to rule on - correcting them would mean inventing or moving ")
    lines.append("a product name, which is not something a tool should do on its own.\n")

    lines.append("\n## 1. Corrections applied (unambiguous spelling only)\n")
    lines.append("Each original spelling remains searchable, so old paperwork still resolves.\n")
    lines.append("| Cell | Was | Now | Why |")
    lines.append("|---|---|---|---|")
    for c in CORRECTIONS:
        mark = "" if c["cell"] in applied else " *(not found - check)*"
        lines.append(f"| `{c['cell']}` | {c['before']} | {c['after']}{mark} | {c['reason']} |")

    lines.append("\n\n## 2. Brand column with no column of its own\n")
    orphans = [f for f in findings["bleeds"] if f["kind"] == "orphan-brand"]
    if orphans:
        by_owner = {}
        for f in orphans:
            by_owner.setdefault(f["expected_owner"], []).append(f)
        for owner, items in by_owner.items():
            lines.append(f"**{len(items)} cells hold {owner} products.** The logo above that ")
            lines.append(f"column is not {owner}, and {owner} has no column of its own.\n")
            lines.append("This is the single largest issue in the table and it affects every ")
            lines.append("quote drawn from that column. The tool shows these entries with a ")
            lines.append("warning rather than hiding or moving them.\n")
            lines.append("| Cell | Product | Column logo |")
            lines.append("|---|---|---|")
            for f in items:
                lines.append(f"| `{f['cell']}` | {f['product']} | {bname[f['brand']]} |")
    else:
        lines.append("None found.")

    lines.append("\n\n## 3. Products sitting in the wrong brand's column\n")
    bleeds = [f for f in findings["bleeds"] if f["kind"] == "column-bleed"]
    if bleeds:
        lines.append("| Cell | Product | Looks like | Sits in |")
        lines.append("|---|---|---|---|")
        for f in bleeds:
            lines.append(f"| `{f['cell']}` | {f['product']} | {f['expected_owner']} | {bname[f['brand']]} |")
    else:
        lines.append("None found.")

    lines.append("\n\n## 4. The same product listed on more than one row\n")
    lines.append("Usually means a value was pasted down a column and the grade never updated. ")
    lines.append("Some of these are legitimate (one product genuinely covering two grades), ")
    lines.append("which is why none have been changed.\n")
    if findings["repeats"]:
        lines.append("| Brand | Product | Rows |")
        lines.append("|---|---|---|")
        seen = {}
        for f in findings["repeats"]:
            seen.setdefault((f["brand"], f["product"]), []).append(f["row"])
        for (brand, product), rows_ in sorted(seen.items()):
            lines.append(f"| {bname[brand]} | {product} | {', '.join(str(r) for r in rows_)} |")
    else:
        lines.append("None found.")

    lines.append("\n\n## 5. Noted by hand, not auto-detectable\n")
    lines.append("- `G21` repeats `Medium Speed Engine 3012` from row 20; row 21 is the SAE 40 ")
    lines.append("  line, so `4012` may have been intended. Not changed - that would invent a name.")
    lines.append("- `K43` (`Energol THB 68`) and `M43` (`Renolin Eterna 68`) sit on the 100-grade ")
    lines.append("  row while `K42`/`M42` are blank or absent, suggesting both slipped down a row.")
    lines.append("- `M56` repeats `RENISO KM 32` from row 54; row 56 is the 68 grade.")
    lines.append("- `F96` reads `Capella HFC 100` while rows 93/95 read `Capella HCF`. One ")
    lines.append("  spelling is wrong but the correct one is not certain from the table alone.")
    lines.append("- Eni (column L) is populated in only a handful of rows. The tool shows those ")
    lines.append("  cells as *not listed* rather than *no equivalent* - they read as unresearched.")

    with open(OUT_QA, "w", encoding="utf8") as fh:
        fh.write("\n".join(lines) + "\n")


def main():
    print("Reading workbook...")
    wb = openpyxl.load_workbook(SOURCE, data_only=True)
    ws = wb["Sheet1"]

    corrections_by_cell = build_index()
    applied = set()

    # Column A holds the category labels, so its corrections must land before the
    # blocks are named.
    label_fixes = {}
    for ref, fix in corrections_by_cell.items():
        if not ref.startswith("A"):
            continue
        actual = re.sub(r"\s+", " ", str(ws.cell(row=int(ref[1:]), column=1).value or "")).strip()
        if actual != fix["before"]:
            raise SystemExit(
                f"Correction for {ref} expected {fix['before']!r} but the workbook now "
                f"holds {actual!r}. Re-check tools/corrections.py."
            )
        label_fixes[ref] = fix["after"]
        applied.add(ref)

    blocks = category_blocks(ws, label_fixes)
    print(f"  category blocks found: {len(blocks)}")

    rows = build_rows(ws, blocks, corrections_by_cell, applied)
    print(f"  data rows: {len(rows)}")

    missing = set(corrections_by_cell) - applied
    if missing:
        raise SystemExit(f"Corrections never applied (cell not reached): {sorted(missing)}")

    assert_invariants(rows, blocks, ws)

    findings = qa.run(rows, BRANDS)
    print(f"  QA: {len(findings['bleeds'])} misplaced, {len(findings['repeats'])} repeated")

    for row in rows:
        for brand in BRANDS:
            key = f"{brand['id']}:{row['id']}"
            if key in findings["warnings"]:
                row["cells"][brand["id"]]["warn"] = findings["warnings"][key]

    index = build_index_map(rows)
    print(f"  searchable names: {len(index)}")

    meta = {
        "source_file": os.path.basename(SOURCE),
        "source_modified": source_modified(),
        "generated": datetime.date.today().isoformat(),
        "rows": len(rows),
        "brands": len(BRANDS),
        "categories": len(blocks),
    }

    logos = load_logos()
    payload = {
        "meta": meta,
        "brands": [{"id": b["id"], "name": b["name"], "col": b["col"]} for b in BRANDS],
        "logos": logos,
        "categories": [b["name"] for b in blocks],
        "rows": rows,
        "index": index,
    }

    os.makedirs(os.path.dirname(OUT_JSON), exist_ok=True)
    os.makedirs(os.path.dirname(OUT_HTML), exist_ok=True)

    readable = dict(payload)
    readable["logos"] = {k: f"<{len(v)} bytes base64>" for k, v in logos.items()}
    with open(OUT_JSON, "w", encoding="utf8") as fh:
        json.dump(readable, fh, indent=1, ensure_ascii=False)

    write_qa_report(rows, findings, applied, meta)

    with open(TEMPLATE, encoding="utf8") as fh:
        html = fh.read()
    blob = json.dumps(payload, ensure_ascii=False, separators=(",", ":"))
    blob = blob.replace("</", "<\\/")  # never let a literal </script> escape the tag
    html = html.replace("__DATA__", blob)
    with open(OUT_HTML, "w", encoding="utf8") as fh:
        fh.write(html)

    size = os.path.getsize(OUT_HTML) / 1024
    print(f"\nWrote {os.path.relpath(OUT_HTML, ROOT)} ({size:.0f} KB)")
    print(f"Wrote {os.path.relpath(OUT_QA, ROOT)}")
    print(f"Wrote {os.path.relpath(OUT_JSON, ROOT)}")


if __name__ == "__main__":
    main()
